# Report Generator module for creating Excel audit reports

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

class ReportGenerator:
    """
    Generates professional Excel audit reports for store signage audits.
    Handles data transformation, formatting, and export.
    Reports present AI findings for human auditor review.
    """
    
    # Output filename constant
    OUTPUT_FILENAME = "signage_audit_report.xlsx"
    
    # Column mapping from results dictionary to Excel columns
    COLUMN_MAPPING = {
        'store_id': 'Store Code',
        'filename': 'Image File',
        'validation_passed': 'Image Quality',
        'signage_visible': 'Signage Visible',
        'light_status': 'Light Status',
        'partial_illumination': 'Partial Letters',
        'issue_detected': 'Issue Detected',
        'issue_type': 'Issue Type',
        'issue_confidence': 'AI Confidence',
        'manual_review_required': 'Manual Review'
    }
    
    def __init__(self):
        """Initialize the Report Generator."""
        pass
    
    def generate_excel_report(self, results, output_path=None):
        """
        Generate a professional Excel audit report from store audit findings.
        
        Reports AI-identified issues for human auditor review and approval.
        
        Args:
            results (list): List of dictionaries with audit findings:
                - store_id: Store identifier
                - filename: Image filename
                - is_readable: Image readable (bool)
                - resolution_ok: Resolution meets minimum (bool)
                - is_blurry: Image is blurry (bool)
                - validation_passed: Overall validation status (bool)
                - signage_visible: Signage visible in image (bool)
                - light_status: Light status string ("ON", "OFF", "UNCLEAR")
                - partial_illumination: Partial letter illumination (bool)
                - issue_detected: Any issues detected by AI (bool)
                - issue_type: Type of issue (str)
                - issue_description: Issue description (str)
                - issue_confidence: AI confidence in findings (float)
                - manual_review_required: Requires human review (bool)
            
            output_path (str, optional): Path to save Excel file.
                If None, saves as 'signage_audit_report.xlsx' in current directory.
        
        Returns:
            str: Path to generated Excel file
        """
        try:
            # Set default output path if not provided
            if output_path is None:
                output_path = os.path.join(os.getcwd(), self.OUTPUT_FILENAME)
            
            # Convert to absolute path
            output_path = os.path.abspath(output_path)
            
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # If file exists and is locked, try to remove it
            if os.path.exists(output_path):
                try:
                    os.remove(output_path)
                except (OSError, PermissionError):
                    # File is locked, try alternative approach
                    pass
            
            # Transform results into structured format
            transformed_data = self._transform_results(results)
            
            # Create DataFrame with required columns
            df = pd.DataFrame(transformed_data)
            
            # Ensure columns are in correct order
            column_order = [
                'Store Code',
                'Image File',
                'Image Quality',
                'Signage Visible',
                'Light Status',
                'Partial Letters',
                'Issue Detected',
                'Issue Type',
                'AI Confidence',
                'Manual Review'
            ]
            
            # Reorder columns (only include those present in dataframe)
            available_columns = [col for col in column_order if col in df.columns]
            df = df[available_columns]
            
            # Save base Excel file
            df.to_excel(output_path, index=False, sheet_name='Audit Report')
            
            # Apply professional formatting
            self._apply_formatting(output_path)
            
            print(f"[OK] Audit report generated successfully: {output_path}")
            print(f"  Total stores audited: {len(df)}")
            issues_detected = len(df[df['Issue Detected'] == 'Yes'])
            print(f"  Issues detected: {issues_detected}")
            manual_review = len(df[df['Manual Review'] == 'Yes'])
            print(f"  Require manual review: {manual_review}")
            
            return output_path
        
        except Exception as e:
            print(f"[X] Error generating report: {str(e)}")
            raise
    
    def _transform_results(self, results):
        """
        Transform raw audit findings into structured format for Excel export.
        
        Converts validation, AI analysis, and audit findings into user-friendly format.
        Converts boolean values to "Yes"/"No" and formats strings appropriately.
        
        Args:
            results (list): Raw audit findings from pipeline
        
        Returns:
            list: List of transformed dictionaries ready for DataFrame
        """
        transformed = []
        
        for result in results:
            # Transform validation result to user-friendly format
            image_quality = "Valid" if result.get('validation_passed', False) else "Invalid"
            signage_visible = "Yes" if result.get('signage_visible', False) else "No"
            partial_letters = "Yes" if result.get('partial_illumination', False) else "No"
            issue_detected = "Yes" if result.get('issue_detected', False) else "No"
            manual_review = "Yes" if result.get('manual_review_required', False) else "No"
            
            # Get light status (handle multiple possible formats)
            light_status = str(result.get('light_status', 'UNCLEAR')).upper()
            
            # Create transformed row
            transformed_row = {
                'Store Code': result.get('store_id', 'N/A'),
                'Image File': result.get('filename', 'N/A'),
                'Image Quality': image_quality,
                'Signage Visible': signage_visible,
                'Light Status': light_status,
                'Partial Letters': partial_letters,
                'Issue Detected': issue_detected,
                'Issue Type': result.get('issue_type', 'unknown'),
                'AI Confidence': f"{result.get('issue_confidence', 0.0):.2f}",
                'Manual Review': manual_review
            }
            
            transformed.append(transformed_row)
        
        return transformed
    
    def _apply_formatting(self, output_path):
        """
        Apply professional formatting to Excel workbook.
        
        Includes:
        - Header row styling (bold, filled background)
        - Column width adjustment for readability
        - Issue type color coding (Red=Issue Detected, Green=No Issue)
        - Borders and alignment
        
        Args:
            output_path (str): Path to Excel file to format
        """
        try:
            # Load the workbook
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Define color scheme for issue detection
            color_scheme = {
                'Yes': 'FFC7CE',  # Light red for issues detected
                'No': 'C6EFCE'    # Light green for no issues
            }
            
            # Format header row (row 1)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Set column widths for better readability
            column_widths = {
                'A': 15,  # Store Code
                'B': 20,  # Image File
                'C': 15,  # Image Quality
                'D': 15,  # Signage Visible
                'E': 15,  # Light Status
                'F': 15,  # Partial Letters
                'G': 15,  # Issue Detected
                'H': 18,  # Issue Type
                'I': 12,  # AI Confidence
                'J': 15   # Manual Review
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.border = thin_border
                
                # Apply color coding to Issue Detected column (column G)
                issue_cell = row[6]  # Index 6 = column G
                issue_value = issue_cell.value
                
                if issue_value in color_scheme:
                    issue_cell.fill = PatternFill(
                        start_color=color_scheme[issue_value],
                        end_color=color_scheme[issue_value],
                        fill_type='solid'
                    )
                    issue_cell.font = Font(bold=True)
            
            # Add borders to header row
            for cell in ws[1]:
                cell.border = thin_border
            
            # Freeze panes (keep header visible when scrolling)
            ws.freeze_panes = 'A2'
            
            # Save formatted workbook
            wb.save(output_path)
            
        except Exception as e:
            print(f"Warning: Could not apply formatting to report: {str(e)}")
            # Report generation still succeeds even if formatting fails