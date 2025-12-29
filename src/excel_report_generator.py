# Excel Report Generator - Create audit-ready Excel output
# Transforms classification results into clean, human-readable Excel format

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelReportGenerator:
    """
    Generates professional Excel reports from classification results.
    
    Output columns:
    - store_code: Store identifier
    - issue_type: One of 6 classification categories
    - confidence: AI confidence score (0.0-1.0)
    - manual_review_required: Always 'Yes' (by design)
    - ai_comment: Human-readable explanation
    """
    
    # Column configuration
    COLUMNS = {
        'store_code': 'Store Code',
        'issue_type': 'Issue Type',
        'confidence': 'Confidence',
        'manual_review_required': 'Manual Review',
        'ai_comment': 'AI Comment'
    }
    
    # AI comment templates for each category
    COMMENT_TEMPLATES = {
        'incorrect_image': 'Image is not Lenskart signage or corrupted/unreadable. Cannot process. Request new image.',
        'cropped_image': 'Signage not fully visible in frame or resolution too low. Resubmit with complete signage in view.',
        'light_not_on': 'Signage lights are completely OFF. Verify if this is expected. Schedule repair if needed.',
        'partial_letters_unlit': 'Some individual letters appear dim or poorly illuminated. Inspect and repair defective letter components.',
        'no_issue_detected': 'No issues detected. All lights appear ON with consistent illumination across letters. Confirm and approve.'
    }
    
    # Column widths for readability
    COLUMN_WIDTHS = {
        'A': 15,  # Store Code
        'B': 20,  # Issue Type
        'C': 12,  # Confidence
        'D': 15,  # Manual Review
        'E': 50   # AI Comment
    }
    
    def __init__(self, output_path=None):
        """
        Initialize the Excel report generator.
        
        Args:
            output_path (str, optional): Path to save Excel file.
                If None, defaults to 'audit_report.xlsx' in current directory.
        """
        self.output_path = output_path or 'audit_report.xlsx'
    
    def generate_from_classifications(self, classification_results):
        """
        Generate Excel report from classification results.
        
        Args:
            classification_results (list): List of classification dicts:
                {
                    'store_id': str,
                    'issue_detected': bool,
                    'issue_type': str,
                    'confidence_score': float,
                    'manual_review_required': bool,
                    'reasoning': str (optional),
                    'details': dict (optional)
                }
        
        Returns:
            str: Path to generated Excel file
        """
        # Transform results into report format
        report_data = []
        
        for result in classification_results:
            row = {
                'store_code': result.get('store_id', 'N/A'),
                'issue_type': result.get('issue_type', 'unknown'),
                'confidence': f"{result.get('confidence_score', 0.0):.2f}",
                'manual_review_required': 'Yes',  # Always Yes by design
                'ai_comment': self._generate_comment(
                    result.get('issue_type', 'unknown'),
                    result.get('reasoning', ''),
                    result.get('confidence_score', 0.0),
                    result.get('details', {})
                )
            }
            report_data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(report_data)
        
        # Ensure columns in correct order
        df = df[['store_code', 'issue_type', 'confidence', 'manual_review_required', 'ai_comment']]
        
        # Save to Excel
        df.to_excel(self.output_path, index=False, sheet_name='Audit Report')
        
        # Apply formatting
        self._apply_formatting()
        
        print(f"[OK] Excel report generated: {self.output_path}")
        print(f"  Total records: {len(df)}")
        
        return self.output_path
    
    def _generate_comment(self, issue_type, reasoning, confidence, details):
        """
        Generate human-readable AI comment.
        
        Combines template with specific details for concise, actionable feedback.
        
        Args:
            issue_type (str): One of 6 categories
            reasoning (str): Explanation from classifier
            confidence (float): Confidence score
            details (dict): Additional details
        
        Returns:
            str: Human-readable comment
        """
        # Get base template
        base_comment = self.COMMENT_TEMPLATES.get(
            issue_type,
            f'Issue type: {issue_type}. Review required.'
        )
        
        # Add confidence qualifier if low
        if confidence < 0.6:
            base_comment += f' (Low confidence: {confidence:.0%})'
        
        # Add specific detail if available
        if issue_type == 'light_not_on' and details:
            base_comment = f'Lights confirmed OFF. {base_comment}'
        
        if issue_type == 'partial_letters_unlit' and details:
            base_comment = f'Partial letter illumination detected. {base_comment}'
        
        if issue_type == 'cropped_image' and details:
            width = details.get('width', '?')
            height = details.get('height', '?')
            base_comment += f' (Image: {width}x{height}px)'
        
        if issue_type == 'unclear_image' and confidence < 0.5:
            base_comment = f'Very uncertain - {base_comment.lower()}'
        
        return base_comment.strip()
    
    def _apply_formatting(self):
        """
        Apply professional formatting to Excel workbook.
        
        Includes:
        - Header styling
        - Column widths
        - Text wrapping
        - Color coding by manual review status
        - Borders and alignment
        """
        try:
            wb = load_workbook(self.output_path)
            ws = wb.active
            
            # Define styles
            header_fill = PatternFill(
                start_color='366092',
                end_color='366092',
                fill_type='solid'
            )
            header_font = Font(
                bold=True,
                color='FFFFFF',
                size=11
            )
            header_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Set column widths
            for col_letter, width in self.COLUMN_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width
            
            # Format data rows
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color-code rows by issue type
            color_scheme = {
                'correct': 'C6EFCE',      # Light green
                'warning': 'FFFFCC',      # Light yellow
                'issue': 'FFC7CE'         # Light red
            }
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
                # Get issue type from column B
                issue_type = ws[f'B{row_idx}'].value or ''
                
                # Determine color
                if issue_type == 'no_issue_detected':
                    color = color_scheme['correct']
                elif issue_type in ['unclear_image']:
                    color = color_scheme['warning']
                else:
                    color = color_scheme['issue']
                
                # Apply formatting to all cells in row
                for cell in row:
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    # Text wrapping for comment column (E)
                    if cell.column == 5:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save formatted workbook
            wb.save(self.output_path)
            
        except Exception as e:
            print(f"[!] Warning: Formatting failed: {str(e)}")
            print(f"    Report still generated but may not be fully formatted")
    
    def generate_from_dict_list(self, records):
        """
        Generate Excel report from list of dictionaries.
        
        Simpler version for direct dict input (not classification objects).
        
        Args:
            records (list): List of dicts with keys:
                - store_code (str)
                - issue_type (str)
                - confidence (float 0.0-1.0)
                - comment (str, optional)
        
        Returns:
            str: Path to generated Excel file
        """
        report_data = []
        
        for record in records:
            row = {
                'store_code': record.get('store_code', 'N/A'),
                'issue_type': record.get('issue_type', 'unknown'),
                'confidence': f"{float(record.get('confidence', 0.0)):.2f}",
                'manual_review_required': 'Yes',
                'ai_comment': record.get('comment') or self.COMMENT_TEMPLATES.get(
                    record.get('issue_type', 'unknown'),
                    'Review required.'
                )
            }
            report_data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(report_data)
        df = df[['store_code', 'issue_type', 'confidence', 'manual_review_required', 'ai_comment']]
        
        # Save to Excel
        df.to_excel(self.output_path, index=False, sheet_name='Audit Report')
        
        # Apply formatting
        self._apply_formatting()
        
        print(f"[OK] Excel report generated: {self.output_path}")
        print(f"  Total records: {len(df)}")
        
        return self.output_path


# ============================================================================
# SIMPLE FUNCTION INTERFACE
# ============================================================================

def create_audit_report(classification_results, output_path='audit_report.xlsx'):
    """
    Simple function to create audit report from classification results.
    
    Args:
        classification_results (list): List of classification dicts
        output_path (str): Where to save Excel file
    
    Returns:
        str: Path to generated Excel file
    
    Example:
        >>> results = [
        ...     {
        ...         'store_id': 'STORE001',
        ...         'issue_type': 'lighting_off',
        ...         'confidence_score': 0.92,
        ...         'manual_review_required': True,
        ...         'reasoning': 'Lights are OFF'
        ...     }
        ... ]
        >>> create_audit_report(results, 'my_report.xlsx')
        '[OK] Excel report generated: my_report.xlsx'
    """
    generator = ExcelReportGenerator(output_path)
    return generator.generate_from_classifications(classification_results)


# ============================================================================
# USAGE EXAMPLES
# ============================================================================

if __name__ == '__main__':
    
    # Example 1: From classification results
    print("Example 1: Generate from classification results")
    print("-" * 70)
    
    sample_results = [
        {
            'store_id': 'LKST1001',
            'issue_type': 'lighting_off',
            'confidence_score': 0.92,
            'manual_review_required': True,
            'reasoning': 'Signage lights are confirmed OFF',
            'details': {}
        },
        {
            'store_id': 'LKST1002',
            'issue_type': 'no_issue_detected',
            'confidence_score': 0.95,
            'manual_review_required': True,
            'reasoning': 'All checks passed',
            'details': {}
        },
        {
            'store_id': 'LKST1003',
            'issue_type': 'cropped_image',
            'confidence_score': 0.90,
            'manual_review_required': True,
            'reasoning': 'Signage not fully visible',
            'details': {'width': 80, 'height': 60}
        },
        {
            'store_id': 'LKST1004',
            'issue_type': 'unclear_image',
            'confidence_score': 0.45,
            'manual_review_required': True,
            'reasoning': 'Low AI confidence',
            'details': {}
        },
        {
            'store_id': 'LKST1005',
            'issue_type': 'partial_lighting',
            'confidence_score': 0.78,
            'manual_review_required': True,
            'reasoning': 'Some letters partially illuminated',
            'details': {}
        }
    ]
    
    # Generate report
    generator = ExcelReportGenerator('example_audit_report.xlsx')
    output_file = generator.generate_from_classifications(sample_results)
    print(f"Report saved to: {output_file}\n")
    
    # Example 2: Simple dict interface
    print("Example 2: Generate from simple dict list")
    print("-" * 70)
    
    simple_records = [
        {
            'store_code': 'STORE001',
            'issue_type': 'lighting_off',
            'confidence': 0.92
        },
        {
            'store_code': 'STORE002',
            'issue_type': 'no_issue_detected',
            'confidence': 0.95
        }
    ]
    
    output_file2 = create_audit_report(simple_records, 'simple_report.xlsx')
    print(f"Report saved to: {output_file2}\n")
    
    # Example 3: Show DataFrame output
    print("Example 3: Display DataFrame preview")
    print("-" * 70)
    
    df_data = []
    for result in sample_results:
        generator = ExcelReportGenerator()
        comment = generator._generate_comment(
            result['issue_type'],
            result['reasoning'],
            result['confidence_score'],
            result['details']
        )
        df_data.append({
            'store_code': result['store_id'],
            'issue_type': result['issue_type'],
            'confidence': f"{result['confidence_score']:.2f}",
            'manual_review_required': 'Yes',
            'ai_comment': comment
        })
    
    df = pd.DataFrame(df_data)
    print("\nDataFrame Preview:")
    print(df.to_string(index=False))
    print(f"\nTotal records: {len(df)}")
